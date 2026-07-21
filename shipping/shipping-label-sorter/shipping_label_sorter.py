#!/usr/bin/env python3
"""
Shipping Label Sorter & 4x6 Thermal Converter

This script scans an input directory for PDF files and ZIP files, extracts ZIP archives,
detects courier partners page-by-page, and splits multi-page PDFs.
For Blue Dart labels, it crops out empty spaces/margins, converts the content to grayscale,
scales it to fit a 4x6 inch portrait thermal page (288 x 432 points), clears the old logo,
overlays assets/logo.png, and moves/shortens weight/dimensions.
Amazon labels are copied/split exactly as-is. Other couriers are skipped.
Output files are saved in the output directory named by Order ID and courier name.
"""

import os
import sys
import argparse
import logging
import traceback
import shutil
import re
import zipfile
from pathlib import Path
try:
    import fitz  # PyMuPDF
except ModuleNotFoundError:
    print("Error: PyMuPDF ('fitz') module not found.", file=sys.stderr)
    print("Please activate the virtual environment first:", file=sys.stderr)
    print("  source .venv/bin/activate", file=sys.stderr)
    print("Or run the script using the virtual environment's python:", file=sys.stderr)
    print("  .venv/bin/python shipping_label_sorter.py", file=sys.stderr)
    sys.exit(1)

import datetime
import subprocess

try:
    import openpyxl
    from openpyxl import Workbook
except ModuleNotFoundError:
    print("Error: openpyxl module not found.", file=sys.stderr)
    print("Please install requirements or activate the virtual environment:", file=sys.stderr)
    print("  .venv/bin/pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)

# Configuration settings
INPUT_DIR = "input"
OUTPUT_DIR = "output"
TEMP_DIR = "temp_extracted"
ASSETS_DIR = "assets"
LOGO_PATH = "assets/logo.png"

THERMAL_WIDTH = 288
THERMAL_HEIGHT = 432

REMOVE_OLD_LOGO = True
ADD_NEW_LOGO = True
DEBUG_LOGO_BOX = False

# Coordinates and details (fine-tuned for Blue Dart thermal labels)
OLD_LOGO_CLEAR_X = 145
OLD_LOGO_CLEAR_Y = 15
OLD_LOGO_CLEAR_WIDTH = 135
OLD_LOGO_CLEAR_HEIGHT = 100

LOGO_X = 165
LOGO_Y = 30
LOGO_WIDTH = 100
LOGO_VERIFY_DPI = 144
LOGO_MIN_DARK_PIXEL_RATIO = 0.02
LOGO_INSERT_ATTEMPTS = 2

# Right-side AWB detail block rewrite (adds Wt/Dim line under Routing Code)
REWRITE_RIGHT_AWB_DETAILS = True
DEBUG_RIGHT_DETAILS_BOX = False

RIGHT_DETAILS_CLEAR_X = 150
RIGHT_DETAILS_CLEAR_Y = 178
RIGHT_DETAILS_CLEAR_WIDTH = 130
RIGHT_DETAILS_CLEAR_HEIGHT = 48

RIGHT_DETAILS_X = 152
RIGHT_DETAILS_Y = 190
RIGHT_DETAILS_LINE_HEIGHT = 10
RIGHT_DETAILS_FONT_SIZE = 7

# Final left-side Dimensions/Weight removal (applied AFTER all other edits)
REMOVE_LEFT_ORIGINAL_WEIGHT_DIMENSIONS_AFTER_ALL_EDITS = True
DEBUG_LEFT_OVERLAY_BOXES = False

# FINAL APPROVED LEFT-SIDE DIMENSIONS/WEIGHT OVERLAY COORDINATES — DO NOT CHANGE
LEFT_DIMENSIONS_CLEAR_X0 = 7
LEFT_DIMENSIONS_CLEAR_Y0 = 119
LEFT_DIMENSIONS_CLEAR_X1 = 140
LEFT_DIMENSIONS_CLEAR_Y1 = 132

LEFT_WEIGHT_CLEAR_X0 = 7
LEFT_WEIGHT_CLEAR_Y0 = 139
LEFT_WEIGHT_CLEAR_X1 = 105
LEFT_WEIGHT_CLEAR_Y1 = 152

# Xpressbees final divider correction
FIX_XPRESSBEES_SHIPPED_BY_DIVIDER = True

XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_X0 = 6
XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_Y0 = 190
XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_X1 = 150
XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_Y1 = 196

XPRESSBEES_SHIPPED_BY_DIVIDER_Y = 226
XPRESSBEES_SHIPPED_BY_DIVIDER_X0 = 5
XPRESSBEES_SHIPPED_BY_DIVIDER_X1 = 283
XPRESSBEES_SHIPPED_BY_DIVIDER_WIDTH = 0.8

DEBUG_XPRESSBEES_DIVIDER_FIX = False

# Print settings
ENABLE_PRINT_PROMPT = True
THERMAL_PRINTER_NAME = ""

# Import modular courier rules registry
from courier_rules import COURIER_MAP


def setup_logging(output_dir: Path) -> logging.Logger:
    """
    Sets up logging to print to the console and write to a log file
    located in the output directory.
    """
    log_format = "%(asctime)s [%(levelname)s] %(message)s"
    log_file = output_dir / "converter.log"

    logger = logging.getLogger("thermal_sorter")
    logger.setLevel(logging.INFO)

    # Clear existing handlers to prevent duplicate logs if re-initialized
    if logger.handlers:
        logger.handlers.clear()

    # File Handler
    try:
        file_handler = logging.FileHandler(log_file, encoding="utf-8")
        file_handler.setFormatter(logging.Formatter(log_format))
        logger.addHandler(file_handler)
    except Exception as e:
        print(f"Warning: Could not create log file at {log_file}: {e}", file=sys.stderr)

    # Console Handler
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setFormatter(logging.Formatter(log_format))
    logger.addHandler(console_handler)

    return logger


def extract_text_with_fallback(page, logger: logging.Logger) -> str:
    """
    Extracts text from a page using native text extraction,
    falling back to OCR if native text is empty and OCR is available.
    """
    text = page.get_text().strip()
    if not text:
        try:
            # Fallback to OCR if native text is empty
            ocr_tp = page.get_textpage_ocr(language="eng", dpi=150)
            text = page.get_text(textpage=ocr_tp).strip()
        except Exception as e:
            logger.debug(f"OCR fallback failed or not available: {e}")
    return text


def extract_weight_dimensions(text: str, default_val: str) -> str:
    """
    Extracts weight and dimensions from text dynamically and returns a shortened line.
    Falls back to default_val if parsing fails or no matches are found.
    """
    weight_match = re.search(r"(?:weight|wt)\s*[:\-]?\s*([\d\.]+)\s*(?:kg|grams|g)?", text, re.IGNORECASE)
    dim_match = re.search(r"(?:dimensions|dims|dim)\s*[:\-]?\s*([^\n]+)", text, re.IGNORECASE)
    
    if not weight_match and not dim_match:
        return default_val
        
    weight = "0.10 kg"
    if weight_match:
        weight = f"{weight_match.group(1).strip()} kg"
        
    dimensions = "30*20*1 cm"
    if dim_match:
        dim_text = dim_match.group(1).strip()
        # Clean up dim_text: remove (cm), cm, spaces, convert 30.00*20.00*1.00 to 30*20*1
        dim_clean = re.sub(r"\(.*?\)", "", dim_text)
        dim_clean = re.sub(r"[a-zA-Z]", "", dim_clean)
        dim_clean = re.sub(r"[\s\*xX]+", "*", dim_clean).strip("*")
        parts = dim_clean.split("*")
        if len(parts) >= 3:
            try:
                formatted_parts = []
                for p in parts[:3]:
                    val = float(p.strip())
                    if val.is_integer():
                        formatted_parts.append(str(int(val)))
                    else:
                        formatted_parts.append(f"{val:.2f}".rstrip("0").rstrip("."))
                dimensions = "*".join(formatted_parts) + " cm"
            except Exception:
                dimensions = dim_clean + " cm"
                
    return f"Wt: {weight} | Dim: {dimensions}"


def extract_xpressbees_right_details(text: str) -> dict:
    awb_match = re.search(r"awb\s*#?\s*[:\-]?\s*(\d+)", text, re.IGNORECASE)
    awb = awb_match.group(1).strip() if awb_match else "N/A"

    routing_codes = [
        match.strip()
        for match in re.findall(r"Routing\s*Code\s*:\s*(.+)", text, re.IGNORECASE)
        if match.strip()
    ]
    routing_code_1 = routing_codes[0] if routing_codes else "N/A"
    routing_code_2 = routing_codes[1] if len(routing_codes) > 1 else "NA"

    weight_match = re.search(r"(?:weight|wt)\s*[:\-]?\s*([\d\.]+)\s*(?:kg|grams|g)?", text, re.IGNORECASE)
    weight = f"{weight_match.group(1).strip()} kg" if weight_match else "0.10 kg"

    dimensions = "30*20*1 cm"
    dim_match = re.search(r"(?:dimensions|dims|dim)\s*[:\-]?\s*([^\n]+)", text, re.IGNORECASE)
    if dim_match:
        dim_text = dim_match.group(1).strip()
        dim_clean = re.sub(r"\(.*?\)", "", dim_text)
        dim_clean = re.sub(r"[a-zA-Z]", "", dim_clean)
        dim_clean = re.sub(r"[\s\*xX]+", "*", dim_clean).strip("*")
        parts = dim_clean.split("*")
        if len(parts) >= 3:
            try:
                formatted_parts = []
                for p in parts[:3]:
                    val = float(p.strip())
                    if val.is_integer():
                        formatted_parts.append(str(int(val)))
                    else:
                        formatted_parts.append(f"{val:.2f}".rstrip("0").rstrip("."))
                dimensions = "*".join(formatted_parts) + " cm"
            except Exception:
                dimensions = dim_clean + " cm"

    return {
        "awb": awb,
        "routing_code_1": routing_code_1,
        "routing_code_2": routing_code_2,
        "weight": weight,
        "dimensions": dimensions,
    }


def flatten_pixmap_on_white(pix: fitz.Pixmap) -> fitz.Pixmap:
    """
    Returns an RGB pixmap with transparent pixels composited onto white.
    This avoids PDF soft-mask rendering issues in viewers and print drivers.
    """
    if pix.colorspace is None or pix.colorspace.n != 3:
        pix = fitz.Pixmap(fitz.csRGB, pix)

    if not pix.alpha:
        return pix

    samples = pix.samples
    rgb_samples = bytearray(pix.width * pix.height * 3)
    src_n = pix.n

    for y in range(pix.height):
        src_row = y * pix.stride
        dst_row = y * pix.width * 3
        for x in range(pix.width):
            src_i = src_row + x * src_n
            dst_i = dst_row + x * 3
            alpha = samples[src_i + src_n - 1]
            inv_alpha = 255 - alpha

            rgb_samples[dst_i] = (samples[src_i] * alpha + 255 * inv_alpha + 127) // 255
            rgb_samples[dst_i + 1] = (samples[src_i + 1] * alpha + 255 * inv_alpha + 127) // 255
            rgb_samples[dst_i + 2] = (samples[src_i + 2] * alpha + 255 * inv_alpha + 127) // 255

    return fitz.Pixmap(fitz.csRGB, pix.width, pix.height, bytes(rgb_samples), 0)


def load_logo_pixmap(logo_path: Path) -> fitz.Pixmap:
    """
    Loads the logo as a plain RGB pixmap with no alpha channel.
    """
    raw_logo = fitz.Pixmap(str(logo_path))
    return flatten_pixmap_on_white(raw_logo)


def get_logo_rect(logo_pix: fitz.Pixmap) -> fitz.Rect:
    aspect_ratio = logo_pix.height / logo_pix.width
    logo_height = LOGO_WIDTH * aspect_ratio
    return fitz.Rect(LOGO_X, LOGO_Y, LOGO_X + LOGO_WIDTH, LOGO_Y + logo_height)


def logo_dark_pixel_ratio(page: fitz.Page, logo_rect: fitz.Rect) -> float:
    """
    Renders just the logo area and returns the ratio of dark pixels.
    A blank or failed logo render has a ratio near zero.
    """
    pix = page.get_pixmap(dpi=LOGO_VERIFY_DPI, colorspace=fitz.csRGB, clip=logo_rect)
    total = pix.width * pix.height
    if total <= 0:
        return 0.0

    samples = pix.samples
    dark_pixels = 0
    for y in range(pix.height):
        row = y * pix.stride
        for x in range(pix.width):
            i = row + x * pix.n
            if min(samples[i], samples[i + 1], samples[i + 2]) < 100:
                dark_pixels += 1

    return dark_pixels / total


def insert_verified_logo(out_page: fitz.Page, logger: logging.Logger) -> bool:
    """
    Inserts the replacement logo as the final overlay and verifies that it renders.
    """
    if not ADD_NEW_LOGO:
        return True

    logo_path = Path(LOGO_PATH)
    if not logo_path.exists():
        if REMOVE_OLD_LOGO:
            logger.warning("Logo not found. Processed without new logo.")
        else:
            logger.warning("Logo not found. Processing without logo.")
        return False

    logo_pix = load_logo_pixmap(logo_path)
    logo_rect = get_logo_rect(logo_pix)

    for attempt in range(1, LOGO_INSERT_ATTEMPTS + 1):
        try:
            out_page.draw_rect(logo_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)
            out_page.insert_image(logo_rect, pixmap=logo_pix, keep_proportion=True, overlay=True)
            dark_ratio = logo_dark_pixel_ratio(out_page, logo_rect)
            if dark_ratio >= LOGO_MIN_DARK_PIXEL_RATIO:
                logger.info(f"  Overlayed and verified logo at {logo_rect} (dark pixel ratio {dark_ratio:.3f})")
                return True

            logger.warning(
                f"  Logo render verification failed on attempt {attempt}: "
                f"dark pixel ratio {dark_ratio:.3f}"
            )
        except Exception as e:
            logger.warning(f"  Failed to overlay logo on attempt {attempt}: {e}")

    logger.warning("  Logo overlay could not be verified after retries.")
    return False


def apply_xpressbees_shipped_by_divider_fix(out_page: fitz.Page, logger: logging.Logger) -> None:
    old_line_rect = fitz.Rect(
        XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_X0,
        XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_Y0,
        XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_X1,
        XPRESSBEES_OLD_PARTIAL_LINE_CLEAR_Y1
    )

    if DEBUG_XPRESSBEES_DIVIDER_FIX:
        out_page.draw_rect(old_line_rect, color=(0, 0, 0), fill=None, overlay=True)
    else:
        out_page.draw_rect(old_line_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)

    out_page.draw_line(
        fitz.Point(XPRESSBEES_SHIPPED_BY_DIVIDER_X0, XPRESSBEES_SHIPPED_BY_DIVIDER_Y),
        fitz.Point(XPRESSBEES_SHIPPED_BY_DIVIDER_X1, XPRESSBEES_SHIPPED_BY_DIVIDER_Y),
        color=(0, 0, 0),
        width=XPRESSBEES_SHIPPED_BY_DIVIDER_WIDTH,
        overlay=True
    )
    logger.info("Applied Xpressbees Shipped By divider fix")


def save_document_atomically(out_doc: fitz.Document, output_path: Path) -> None:
    """
    Saves to a hidden temporary file first, then moves it into place.
    This prevents half-written PDFs from appearing in the output folder.
    """
    tmp_path = output_path.with_name(f".{output_path.name}.tmp")
    if tmp_path.exists():
        tmp_path.unlink()

    try:
        out_doc.save(tmp_path, garbage=4, deflate=True)
        out_doc.close()
        tmp_path.replace(output_path)
    except Exception:
        try:
            out_doc.close()
        finally:
            if tmp_path.exists():
                tmp_path.unlink()
        raise


def build_output_filename(order_id: str, courier: str, suffix: str) -> str:
    """
    Builds output filename base using order ID or fallback prefix, courier, and suffix.
    """
    return f"{order_id}_{courier}_{suffix}"


def normalize_courier_slug(courier_text):
    text = (courier_text or "").lower()
    compact = text.replace(" ", "").replace("-", "").replace("_", "")

    if "bluedart" in compact:
        return "bluedart"

    if "xpressbees" in compact or "xpressbee" in compact:
        return "xpressbees"

    if "amazon" in compact:
        return "amazon"

    return "unknown"


def is_bluedart_like_courier(courier_slug):
    return courier_slug in {"bluedart", "xpressbees"}


def get_unique_filename(output_dir: Path, base_name: str, ext: str) -> Path:
    """
    Returns a unique Path in output_dir by appending a counter suffix if the file exists.
    E.g., base_name = "1101_bluedart_4x6", ext = ".pdf"
    """
    file_path = output_dir / f"{base_name}{ext}"
    if not file_path.exists():
        return file_path
        
    counter = 2
    while True:
        file_path = output_dir / f"{base_name}_{counter}{ext}"
        if not file_path.exists():
            return file_path
        counter += 1


def extract_order_id(text: str) -> str:
    """
    Extracts Order ID from label text using case-insensitive regex pattern.
    Cleans it up to keep only letters, numbers, hyphens, and underscores.
    """
    # 1. Try matching invoice slash-based pattern (e.g. Retail00004/1109)
    invoice_pattern = r"\b(?:invoice\s*(?:id|1d))(?!\w)\s*[:\-#]?\s*[a-zA-Z0-9_]+\/([a-zA-Z0-9_\-]+)"
    match = re.search(invoice_pattern, text, re.IGNORECASE)
    if match:
        raw_id = match.group(1).strip()
        cleaned_id = re.sub(r"[^a-zA-Z0-9_\-]", "", raw_id)
        if cleaned_id:
            return cleaned_id

    # 2. Match standard order ID patterns with word boundaries and negative lookahead
    pattern = r"\b(?:seller\s*order\s*id|shopify\s*order\s*id|merchant\s*order\s*id|external\s*order\s*id|reference\s*id|ref\s*id|order\s*number|order\s*id|order\s*no\.?|order\s*#|order)(?!\w)\s*[:\-#]?\s*([a-zA-Z0-9_\-]+)"
    match = re.search(pattern, text, re.IGNORECASE)
    if match:
        raw_id = match.group(1).strip()
        cleaned_id = re.sub(r"[^a-zA-Z0-9_\-]", "", raw_id)
        if cleaned_id:
            return cleaned_id
    return None


def detect_courier_for_page(page, file_path: Path, logger: logging.Logger) -> str:
    """
    Extracts text from a single page and checks for courier keywords.
    Falls back to checking the filename if text extraction is empty or yields no keywords.
    Returns: 'bluedart', 'amazon', 'delhivery', or 'unknown'.
    """
    text = extract_text_with_fallback(page, logger)
    text_lower = text.lower().strip()
    
    # Detection keywords
    amazon_keywords = ["amazon", "amazon shipping", "ats", "amazon transportation"]
    delhivery_keywords = ["delhivery"]
    
    # 1. Try detecting from PDF text content
    courier_slug = normalize_courier_slug(text_lower)
    if courier_slug != "unknown":
        return courier_slug
        
    if any(keyword in text_lower for keyword in amazon_keywords):
        return "amazon"
        
    if any(keyword in text_lower for keyword in delhivery_keywords):
        return "delhivery"
        
    # 2. Fallback to checking the filename
    filename_lower = file_path.name.lower()
    courier_slug = normalize_courier_slug(filename_lower)
    if courier_slug != "unknown":
        return courier_slug
        
    if any(keyword in filename_lower for keyword in amazon_keywords):
        return "amazon"
        
    if any(keyword in filename_lower for keyword in delhivery_keywords):
        return "delhivery"
        
    return "unknown"


def extract_zip(zip_path: Path, temp_dir: Path, logger: logging.Logger) -> list:
    """
    Extracts a zip file to temp_dir/<zip_name> and returns a list of extracted PDF paths.
    """
    extracted_pdfs = []
    target_dir = temp_dir / zip_path.stem
    target_dir.mkdir(parents=True, exist_ok=True)
    
    try:
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(target_dir)
        logger.info(f"Extracted ZIP {zip_path.name} to {target_dir}")
        
        # Recursively find all PDF files
        for p in target_dir.rglob("*.pdf"):
            if p.is_file():
                extracted_pdfs.append(p)
        logger.info(f"Discovered {len(extracted_pdfs)} PDF files inside {zip_path.name}")
    except Exception as e:
        logger.error(f"Failed to extract/read ZIP {zip_path.name}: {e}")
        
    return extracted_pdfs


def get_unique_filename(output_dir: Path, base_name: str, ext: str) -> Path:
    """
    Returns a unique Path in output_dir by appending a counter suffix if the file exists.
    E.g., base_name = "1101_bluedart_4x6", ext = ".pdf"
    """
    file_path = output_dir / f"{base_name}{ext}"
    if not file_path.exists():
        return file_path
        
    counter = 2
    while True:
        file_path = output_dir / f"{base_name}_{counter}{ext}"
        if not file_path.exists():
            return file_path
        counter += 1


def process_bluedart_page(src_page, output_path: Path, page_text: str, logger: logging.Logger, courier_slug: str = "bluedart", order_id: str = "") -> bool:
    """
    Processes a single page visual transformation for Blue Dart courier.
    """
    try:
        rule_module = COURIER_MAP.get("bluedart", COURIER_MAP["unknown"])
        out_doc = fitz.open()
        
        config = {
            "threshold": 250,
            "gap_limit": 150,
            "border_exclude": 80,
            "dpi": 300,
            "padding": 5.0
        }
        
        pix_cropped, crop_rect = rule_module.process_page(src_page, config)
        
        if pix_cropped and crop_rect:
            # Create output 4x6 portrait page
            out_page = out_doc.new_page(width=THERMAL_WIDTH, height=THERMAL_HEIGHT)
            
            # Calculate fit dimensions to fill 4x6 portrait canvas
            w_pt = crop_rect.width
            h_pt = crop_rect.height
            scale = min(THERMAL_WIDTH / w_pt, THERMAL_HEIGHT / h_pt)
            scaled_w = w_pt * scale
            scaled_h = h_pt * scale
            
            margin_x = (THERMAL_WIDTH - scaled_w) / 2.0
            margin_y = (THERMAL_HEIGHT - scaled_h) / 2.0
            
            dest_rect = fitz.Rect(margin_x, margin_y, margin_x + scaled_w, margin_y + scaled_h)
            logger.info(f"  Bbox {crop_rect} scaled by {scale:.4f} to {dest_rect}")
            
            # Insert the high resolution grayscale cropped image
            out_page.insert_image(dest_rect, pixmap=pix_cropped)
            
            # 1. Clear old logo area if enabled
            if REMOVE_OLD_LOGO:
                try:
                    clear_rect = fitz.Rect(
                        OLD_LOGO_CLEAR_X, 
                        OLD_LOGO_CLEAR_Y, 
                        OLD_LOGO_CLEAR_X + OLD_LOGO_CLEAR_WIDTH, 
                        OLD_LOGO_CLEAR_Y + OLD_LOGO_CLEAR_HEIGHT
                    )
                    if DEBUG_LOGO_BOX:
                        # Draw black border, no fill
                        out_page.draw_rect(clear_rect, color=(0, 0, 0), fill=None, overlay=True)
                        logger.info(f"  [DEBUG] Drew logo clear box border at {clear_rect}")
                    else:
                        # Normal white fill
                        out_page.draw_rect(clear_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)
                        logger.info(f"  Cleared old logo area at {clear_rect}")
                except Exception as e:
                    logger.warning(f"  Failed to clear old logo area: {e}")

            # 2. Rewrite right-side AWB detail block with added Wt/Dim line
            if REWRITE_RIGHT_AWB_DETAILS:
                if courier_slug == "xpressbees":
                    logger.info("Applying right-side Wt/Dim block")
                try:
                    import re as _re

                    # Extract values from label text
                    awb_match = _re.search(r"awb\s*#?\s*[:\-]?\s*(\d+)", page_text, _re.IGNORECASE)
                    awb = awb_match.group(1).strip() if awb_match else "N/A"

                    routing_match = _re.search(r"routing\s*code\s*[:\-]?\s*([^\n]+)", page_text, _re.IGNORECASE)
                    routing_code = routing_match.group(1).strip() if routing_match else "N/A"

                    cluster_match = _re.search(r"cluster\s*code\s*[:\-]?\s*([^\n]+)", page_text, _re.IGNORECASE)
                    cluster_code = cluster_match.group(1).strip() if cluster_match else "N/A"

                    # Build shortened Wt/Dim string using existing helper
                    weight_dim_str = extract_weight_dimensions(page_text, "Wt: 0.10 kg | Dim: 30*20*1 cm")
                    xpressbees_details = None
                    if courier_slug == "xpressbees":
                        xpressbees_details = extract_xpressbees_right_details(page_text)
                        awb = xpressbees_details["awb"]
                        routing_code_1 = xpressbees_details["routing_code_1"]
                        routing_code_2 = xpressbees_details["routing_code_2"]
                        weight = xpressbees_details["weight"]
                        dimensions = xpressbees_details["dimensions"]
                        logger.info(f"Xpressbees order_id: {order_id}")
                        logger.info(f"Xpressbees awb: {awb}")
                        logger.info(f"Xpressbees routing_code_1: {routing_code_1}")
                        logger.info(f"Xpressbees routing_code_2: {routing_code_2}")
                        logger.info(f"Xpressbees weight: {weight}")
                        logger.info(f"Xpressbees dimensions: {dimensions}")

                    # Clear the right-side text area (does not touch AWB barcode above)
                    right_clear_rect = fitz.Rect(
                        RIGHT_DETAILS_CLEAR_X,
                        RIGHT_DETAILS_CLEAR_Y,
                        RIGHT_DETAILS_CLEAR_X + RIGHT_DETAILS_CLEAR_WIDTH,
                        RIGHT_DETAILS_CLEAR_Y + RIGHT_DETAILS_CLEAR_HEIGHT
                    )
                    if DEBUG_RIGHT_DETAILS_BOX:
                        out_page.draw_rect(right_clear_rect, color=(0, 0, 0), fill=None, overlay=True, width=0.3)
                        logger.info(f"  [DEBUG] Right details clear box outline at {right_clear_rect}")
                    else:
                        out_page.draw_rect(right_clear_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)

                        # Insert four evenly spaced lines
                        y = RIGHT_DETAILS_Y
                        if courier_slug == "xpressbees":
                            right_lines = [
                                f"Awb: {awb}",
                                f"Routing Code: {routing_code_1}",
                                f"Routing Code: {routing_code_2}",
                                f"Wt: {weight} | Dim: {dimensions}",
                            ]
                            for line in right_lines:
                                out_page.insert_text(fitz.Point(RIGHT_DETAILS_X, y), line, fontsize=RIGHT_DETAILS_FONT_SIZE, color=(0, 0, 0))
                                y += RIGHT_DETAILS_LINE_HEIGHT
                            logger.info("Applied Xpressbees right-side detail rewrite")
                        else:
                            out_page.insert_text(fitz.Point(RIGHT_DETAILS_X, y), f"Awb: {awb}", fontsize=RIGHT_DETAILS_FONT_SIZE, color=(0, 0, 0))
                            y += RIGHT_DETAILS_LINE_HEIGHT
                            out_page.insert_text(fitz.Point(RIGHT_DETAILS_X, y), f"Routing Code: {routing_code}", fontsize=RIGHT_DETAILS_FONT_SIZE, color=(0, 0, 0))
                            y += RIGHT_DETAILS_LINE_HEIGHT
                            out_page.insert_text(fitz.Point(RIGHT_DETAILS_X, y), weight_dim_str, fontsize=RIGHT_DETAILS_FONT_SIZE, color=(0, 0, 0))
                            y += RIGHT_DETAILS_LINE_HEIGHT
                            out_page.insert_text(fitz.Point(RIGHT_DETAILS_X, y), f"Cluster Code: {cluster_code}", fontsize=RIGHT_DETAILS_FONT_SIZE, color=(0, 0, 0))

                            logger.info(f"  Rewrote right-side details: Awb={awb}, Routing={routing_code}, {weight_dim_str}, Cluster={cluster_code}")
                    logger.info("  Right-side AWB detail block rewritten")
                except Exception as e:
                    logger.warning(f"  Failed to rewrite right-side AWB details: {e}")

            # 3. Remove original left-side Dimensions and Weight text
            # 123456789
            if REMOVE_LEFT_ORIGINAL_WEIGHT_DIMENSIONS_AFTER_ALL_EDITS:
                if courier_slug == "xpressbees":
                    logger.info("Applying final left-side Dimensions/Weight overlay")
                try:
                    # Dynamically locate the Dimensions text box on the left side
                    y0_dim = None
                    try:
                        words = src_page.get_text("words")
                        for w in words:
                            x0, y0, x1, y1, word_text = w[0], w[1], w[2], w[3], w[4].lower()
                            if x0 < 150:  # Only look at the left half of the label
                                if "dimension" in word_text:
                                    y0_dim = y0
                                    break
                    except Exception as lookup_err:
                        logger.warning(f"  Could not dynamically locate left-side text: {lookup_err}")

                    # If we found the dynamic top of the Dimensions line, clear that section
                    # and shift the bottom details (COD, Payment, Items) upwards to close the gap
                    same_thermal_layout = (
                        abs(src_page.rect.width - THERMAL_WIDTH) < 1.0
                        and abs(src_page.rect.height - THERMAL_HEIGHT) < 1.0
                        and y0_dim < 228.0
                    )

                    if y0_dim is not None and same_thermal_layout:
                        # 1. Clear the entire section from y0_dim to Y=228 (the horizontal line)
                        clear_rect = fitz.Rect(LEFT_DIMENSIONS_CLEAR_X0, y0_dim - 1.0, 148.0, 228.0)
                        out_page.draw_rect(clear_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)

                        # 2. Shift the bottom details up by 33.3 points using show_pdf_page with clip
                        clip_rect = fitz.Rect(LEFT_DIMENSIONS_CLEAR_X0, y0_dim + 33.3, 148.0, 228.0)
                        dest_rect = fitz.Rect(LEFT_DIMENSIONS_CLEAR_X0, y0_dim, 148.0, 228.0 - 33.3)
                        out_page.show_pdf_page(dest_rect, src_page.parent, pno=src_page.number, clip=clip_rect)
                        logger.info(f"  Applied final left-side Dimensions/Weight removal and shifted bottom details upwards by 33.3 points.")
                        if courier_slug == "xpressbees":
                            logger.info("Applied final left-side Dimensions/Weight overlay")
                    else:
                        # Fallback to static coordinates if not found dynamically (clears without shifting)
                        dim_rect = fitz.Rect(
                            LEFT_DIMENSIONS_CLEAR_X0,
                            LEFT_DIMENSIONS_CLEAR_Y0,
                            LEFT_DIMENSIONS_CLEAR_X1,
                            LEFT_DIMENSIONS_CLEAR_Y1
                        )
                        wt_rect = fitz.Rect(
                            LEFT_WEIGHT_CLEAR_X0,
                            LEFT_WEIGHT_CLEAR_Y0,
                            LEFT_WEIGHT_CLEAR_X1,
                            LEFT_WEIGHT_CLEAR_Y1
                        )
                        out_page.draw_rect(dim_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)
                        out_page.draw_rect(wt_rect, color=(1, 1, 1), fill=(1, 1, 1), overlay=True)
                        logger.info(f"  Fallback: cleared left-side Dimensions/Weight using static coordinates.")
                        if courier_slug == "xpressbees":
                            logger.info("Applied final left-side Dimensions/Weight overlay")
                except Exception as e:
                    logger.warning(f"  Failed to remove/shift left-side Dimensions/Weight: {e}")

            # 4. Add the replacement logo last so later page edits cannot cover it.
            if courier_slug == "xpressbees":
                logger.info("Applying logo replacement")
            insert_verified_logo(out_page, logger)
            if courier_slug == "xpressbees" and FIX_XPRESSBEES_SHIPPED_BY_DIVIDER:
                apply_xpressbees_shipped_by_divider_fix(out_page, logger)
        else:
            logger.warning("  Empty page or no content detected. Inserting empty 4x6 page.")
            out_doc.new_page(width=THERMAL_WIDTH, height=THERMAL_HEIGHT)
            
        save_document_atomically(out_doc, output_path)
        return True
    except Exception as e:
        logger.error(f"Failed to process Blue Dart page: {e}")
        logger.error(traceback.format_exc())
        return False


def order_already_processed(order_id, output_dir=OUTPUT_DIR):
    """
    Returns True if any PDF in output_dir starts with f"{order_id}_".
    """
    if not order_id:
        return False
    output_path = Path(output_dir)
    if not output_path.exists():
        return False
    
    prefix = f"{order_id}_"
    try:
        for item in output_path.iterdir():
            if item.is_file() and item.suffix.lower() == ".pdf":
                name_lower = item.name.lower()
                if name_lower.startswith(prefix.lower()):
                    return True
    except Exception:
        pass
    return False


def append_excel_log(row_data, output_dir=OUTPUT_DIR):
    """
    Appends a row to converter_log.xlsx in the specified output_dir.
    """
    output_path = Path(output_dir)
    excel_path = output_path / "converter_log.xlsx"
    headers = ["timestamp", "source_file", "page_number", "order_id", "courier", "action", "output_file", "status", "message"]
    
    # Ensure directory exists
    output_path.mkdir(parents=True, exist_ok=True)
    
    wb = None
    if excel_path.exists():
        try:
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
        except Exception:
            wb = Workbook()
            sheet = wb.active
            sheet.append(headers)
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(headers)
        
    if isinstance(row_data, dict):
        row_list = [row_data.get(h, "") for h in headers]
    else:
        row_list = list(row_data)
        
    sheet.append(row_list)
    wb.save(excel_path)
    wb.close()


def find_label_by_order_id(order_id, output_dir=OUTPUT_DIR):
    """
    Returns matching PDF path if any file in output_dir starts with f"{order_id}_".
    """
    if not order_id:
        return None
    output_path = Path(output_dir)
    if not output_path.exists():
        return None
    prefix = f"{order_id}_"
    try:
        for item in output_path.iterdir():
            if item.is_file() and item.name.lower().startswith(prefix.lower()) and item.suffix.lower() == ".pdf":
                return item
    except Exception:
        pass
    return None


def print_label_pdf(pdf_path):
    """
    Prints the PDF using lp command with 4x6 portrait settings.
    Logs the outcome to both text and Excel logs.
    """
    pdf_path = Path(pdf_path)
    logger = logging.getLogger("thermal_sorter")
    
    # Extract metadata from filename for logging
    stem_parts = pdf_path.name.split("_")
    if len(stem_parts) >= 2:
        order_id = stem_parts[0]
        courier = stem_parts[1]
    else:
        order_id = pdf_path.stem
        courier = "unknown"
        
    if not pdf_path.exists():
        err_msg = f"PDF file not found: {pdf_path}"
        logger.error(err_msg)
        append_excel_log({
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_file": "",
            "page_number": "",
            "order_id": order_id,
            "courier": courier,
            "action": "print_failed",
            "output_file": pdf_path.name,
            "status": "failed",
            "message": err_msg
        }, output_dir=pdf_path.parent)
        return False

    if THERMAL_PRINTER_NAME:
        cmd = [
            "lp",
            "-d", THERMAL_PRINTER_NAME,
            "-o", "media=Custom.4x6in",
            "-o", "orientation-requested=3",
            str(pdf_path)
        ]
    else:
        cmd = [
            "lp",
            "-o", "media=Custom.4x6in",
            "-o", "orientation-requested=3",
            str(pdf_path)
        ]

    logger.info(f"Printing label: {pdf_path.name} using command: {' '.join(cmd)}")
    
    try:
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
        if result.returncode == 0:
            logger.info(f"Successfully printed: {pdf_path.name}")
            append_excel_log({
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": "",
                "page_number": "",
                "order_id": order_id,
                "courier": courier,
                "action": "printed",
                "output_file": pdf_path.name,
                "status": "success",
                "message": "Sent to printer"
            }, output_dir=pdf_path.parent)
            return True
        else:
            err_msg = f"lp command failed: {result.stderr.strip()}"
            logger.error(err_msg)
            append_excel_log({
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": "",
                "page_number": "",
                "order_id": order_id,
                "courier": courier,
                "action": "print_failed",
                "output_file": pdf_path.name,
                "status": "failed",
                "message": err_msg
            }, output_dir=pdf_path.parent)
            return False
    except Exception as e:
        err_msg = f"Failed to run print command: {e}"
        logger.error(err_msg)
        append_excel_log({
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "source_file": "",
            "page_number": "",
            "order_id": order_id,
            "courier": courier,
            "action": "print_failed",
            "output_file": pdf_path.name,
            "status": "failed",
            "message": err_msg
        }, output_dir=pdf_path.parent)
        return False


def interactive_print_prompt(output_dir=OUTPUT_DIR):
    """
    Repeatedly asks user for Order ID and prints if found in output directory.
    """
    if not ENABLE_PRINT_PROMPT:
        return

    prompt_msg = "\nEnter Order ID to print, or press Enter to exit: "
    while True:
        try:
            order_id_input = input(prompt_msg).strip()
        except (KeyboardInterrupt, EOFError):
            print()
            break
            
        if not order_id_input:
            break
            
        pdf_path = find_label_by_order_id(order_id_input, output_dir)
        if pdf_path:
            print_label_pdf(pdf_path)
        else:
            print(f"No processed label found for Order ID: {order_id_input}")
            
        prompt_msg = "Enter another Order ID to print, or press Enter to exit: "


def main():
    parser = argparse.ArgumentParser(description="Convert and sort shipping labels.")
    parser.add_argument("--input", default=INPUT_DIR, help="Input directory")
    parser.add_argument("--output", default=OUTPUT_DIR, help="Output directory")
    parser.add_argument(
        "--force",
        action="store_true",
        help="Regenerate labels even when an output PDF for the Order ID already exists",
    )
    args = parser.parse_args()

    input_dir = Path(args.input).resolve()
    output_dir = Path(args.output).resolve()
    temp_dir = Path(TEMP_DIR).resolve()
    assets_dir = Path(ASSETS_DIR).resolve()

    # Create directories if missing
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    temp_dir.mkdir(parents=True, exist_ok=True)
    assets_dir.mkdir(parents=True, exist_ok=True)

    # Set up logger
    logger = setup_logging(output_dir)

    logger.info("=" * 60)
    logger.info("PDF Thermal Sorter & Converter Started (Blue Dart & Amazon)")
    logger.info(f"Input directory: {input_dir}")
    logger.info(f"Output directory: {output_dir}")
    logger.info("=" * 60)

    # Scan for input files (PDF and ZIP)
    zip_files = sorted([f for f in input_dir.iterdir() if f.is_file() and f.suffix.lower() == ".zip"])
    direct_pdfs = sorted([f for f in input_dir.iterdir() if f.is_file() and f.suffix.lower() == ".pdf"])

    logger.info(f"Files found in input directory: {len(direct_pdfs) + len(zip_files)}")
    for f in direct_pdfs:
        logger.info(f" - [PDF] {f.name}")
    for f in zip_files:
        logger.info(f" - [ZIP] {f.name}")

    # Build queue of PDFs to process: tuple of (file_path, original_source_name)
    pdf_to_process = []
    for p in direct_pdfs:
        pdf_to_process.append((p, p.name))

    for z in zip_files:
        logger.info(f"Processing ZIP file: {z.name}")
        extracted_pdfs = extract_zip(z, temp_dir, logger)
        for p in extracted_pdfs:
            pdf_to_process.append((p, z.name))

    if not pdf_to_process:
        logger.info("No PDF files found to process. Exiting.")
        return

    success_count = 0
    skipped_count = 0
    failed_count = 0
    output_paths_written_this_run = set()
    processed_order_ids_this_run = set()

    for file_path, source_archive in pdf_to_process:
        logger.info("-" * 40)
        logger.info(f"Processing PDF file: {file_path.name}")
        try:
            src_doc = fitz.open(file_path)
        except Exception as e:
            logger.error(f"Failed to open PDF file {file_path.name}: {e}")
            failed_count += 1
            append_excel_log({
                "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "source_file": file_path.name,
                "page_number": "",
                "order_id": "",
                "courier": "",
                "action": "failed",
                "output_file": "",
                "status": "failed",
                "message": f"Failed to open PDF file: {e}"
            }, output_dir=output_dir)
            continue

        num_pages = len(src_doc)
        for pno in range(num_pages):
            logger.info(f"Processing: {file_path.name} page {pno + 1}")
            try:
                src_page = src_doc[pno]
                page_text = extract_text_with_fallback(src_page, logger)

                # Detect courier
                courier = detect_courier_for_page(src_page, file_path, logger)
                if courier in ["unknown", "undetected"] or not courier:
                    courier = "unknown"
                    logger.info("Courier not detected; using unknown")
                else:
                    if courier == "xpressbees":
                        logger.info("Detected courier: Xpressbees")
                    else:
                        logger.info(f"Detected courier: {courier}")

                # Extract Order ID
                if courier == "amazon":
                    order_id = None
                    # Run diagnostic/debug task for Amazon
                    amazon_page_text = page_text
                    logger.info("========== AMAZON LABEL TEXT DEBUG ==========")
                    logger.info(f"Source file: {file_path}")
                    logger.info(f"Page number: {pno + 1}")
                    logger.info(amazon_page_text[:3000])
                    logger.info("========== END AMAZON LABEL TEXT DEBUG ==========")

                    detected_id = extract_order_id(amazon_page_text)
                    if not detected_id:
                        patterns = [
                            r"Order\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Order\s*Id\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"ORDER\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Order\s*#\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"ORDER\s*#\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Order\s*No\.?\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Order\s*Number\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Seller\s*Order\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Reference\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Ref\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Merchant\s*Order\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"External\s*Order\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)",
                            r"Amazon\s*Order\s*ID\s*[:#-]?\s*([A-Za-z0-9_-]+)"
                        ]
                        for pattern in patterns:
                            match = re.search(pattern, amazon_page_text, re.IGNORECASE)
                            if match:
                                raw_id = match.group(1).strip()
                                cleaned_id = re.sub(r"[^a-zA-Z0-9_\-]", "", raw_id)
                                if cleaned_id:
                                    detected_id = cleaned_id
                                    break
                    
                    if detected_id:
                        logger.info(f"Amazon detected order id: {detected_id}")
                        order_id = detected_id
                    else:
                        logger.warning("Amazon order id could not be detected from extracted page text.")

                    print("Amazon label found:")
                    print(f"Source: {file_path.name}")
                    print(f"Page: {pno + 1}")
                    print(f"Detected Order ID: {detected_id if detected_id else 'NOT FOUND'}")
                else:
                    order_id = extract_order_id(page_text)

                if not order_id:
                    logger.info("Order ID not found")
                    # Fallback naming prefix: unknown_order_<sourcefilename>_p<page_number>
                    order_id_fallback = f"unknown_order_{file_path.stem}_p{pno + 1}"
                else:
                    logger.info(f"Order ID: {order_id}")
                    order_id_fallback = order_id

                # Before processing, check if Order ID already exists in output folder
                order_id_key = order_id.lower() if order_id else None
                if order_id_key and order_id_key in processed_order_ids_this_run:
                    logger.info(f"Skipped duplicate Order ID in current run: {order_id}")
                    skipped_count += 1
                    append_excel_log({
                        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "source_file": file_path.name,
                        "page_number": pno + 1,
                        "order_id": order_id,
                        "courier": courier,
                        "action": "skipped_duplicate_in_current_run",
                        "output_file": "",
                        "status": "skipped",
                        "message": f"Skipped duplicate Order ID in current run: {order_id}"
                    }, output_dir=output_dir)
                    continue

                if order_id and not args.force and order_already_processed(order_id, output_dir):
                    logger.info(f"Skipped already processed Order ID: {order_id}")
                    skipped_count += 1
                    append_excel_log({
                        "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                        "source_file": file_path.name,
                        "page_number": pno + 1,
                        "order_id": order_id,
                        "courier": courier,
                        "action": "skipped_already_processed",
                        "output_file": "",
                        "status": "skipped",
                        "message": f"Skipped already processed Order ID: {order_id}"
                    }, output_dir=output_dir)
                    continue

                # Define output filename base using shared build_output_filename function
                base_out_name = build_output_filename(order_id_fallback, courier, "4x6")
                forced_out_path = output_dir / f"{base_out_name}.pdf"
                if args.force and forced_out_path not in output_paths_written_this_run:
                    out_path = forced_out_path
                else:
                    out_path = get_unique_filename(output_dir, base_out_name, ".pdf")
                output_paths_written_this_run.add(out_path)

                # Save based on courier
                if is_bluedart_like_courier(courier):
                    if courier == "xpressbees":
                        logger.info("Using Blue Dart-like processing pipeline for Xpressbees")
                    success = process_bluedart_page(src_page, out_path, page_text, logger, courier, order_id if order_id else "")
                    if success:
                        success_count += 1
                        if order_id_key:
                            processed_order_ids_this_run.add(order_id_key)
                        logger.info(f"Saved: output/{out_path.name}")
                        append_excel_log({
                            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "source_file": file_path.name,
                            "page_number": pno + 1,
                            "order_id": order_id if order_id else "",
                            "courier": courier,
                            "action": "processed",
                            "output_file": out_path.name,
                            "status": "success",
                            "message": f"Saved: output/{out_path.name}"
                        }, output_dir=output_dir)
                    else:
                        failed_count += 1
                        append_excel_log({
                            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "source_file": file_path.name,
                            "page_number": pno + 1,
                            "order_id": order_id if order_id else "",
                            "courier": courier,
                            "action": "failed",
                            "output_file": "",
                            "status": "failed",
                            "message": "Failed to process Blue Dart page"
                        }, output_dir=output_dir)
                else:
                    # Amazon, other couriers, or unknown: plain 4x6 portrait vector scale
                    try:
                        out_doc = fitz.open()
                        out_page = out_doc.new_page(width=THERMAL_WIDTH, height=THERMAL_HEIGHT)
                        
                        w_pt = src_page.rect.width
                        h_pt = src_page.rect.height
                        scale = min(THERMAL_WIDTH / w_pt, THERMAL_HEIGHT / h_pt)
                        scaled_w = w_pt * scale
                        scaled_h = h_pt * scale
                        
                        margin_x = (THERMAL_WIDTH - scaled_w) / 2.0
                        margin_y = (THERMAL_HEIGHT - scaled_h) / 2.0
                        
                        dest_rect = fitz.Rect(margin_x, margin_y, margin_x + scaled_w, margin_y + scaled_h)
                        out_page.show_pdf_page(dest_rect, src_doc, pno=pno)
                        
                        save_document_atomically(out_doc, out_path)
                        success_count += 1
                        if order_id_key:
                            processed_order_ids_this_run.add(order_id_key)
                        logger.info(f"Saved: output/{out_path.name}")
                        append_excel_log({
                            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "source_file": file_path.name,
                            "page_number": pno + 1,
                            "order_id": order_id if order_id else "",
                            "courier": courier,
                            "action": "copied",
                            "output_file": out_path.name,
                            "status": "success",
                            "message": f"Saved: output/{out_path.name}"
                        }, output_dir=output_dir)
                    except Exception as e:
                        logger.error(f"Failed to process {courier} page {pno + 1}: {e}")
                        failed_count += 1
                        append_excel_log({
                            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "source_file": file_path.name,
                            "page_number": pno + 1,
                            "order_id": order_id if order_id else "",
                            "courier": courier,
                            "action": "failed",
                            "output_file": "",
                            "status": "failed",
                            "message": f"Failed to process {courier} page {pno + 1}: {e}"
                        }, output_dir=output_dir)

            except Exception as e:
                logger.error(f"Error processing page {pno + 1} of {file_path.name}: {e}")
                logger.error(traceback.format_exc())
                failed_count += 1
                append_excel_log({
                    "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "source_file": file_path.name,
                    "page_number": pno + 1,
                    "order_id": order_id if 'order_id' in locals() and order_id else "",
                    "courier": courier if 'courier' in locals() and courier else "",
                    "action": "failed",
                    "output_file": "",
                    "status": "failed",
                    "message": f"Error processing page {pno + 1}: {e}"
                }, output_dir=output_dir)

        src_doc.close()

    # Final summary logs
    logger.info("=" * 60)
    logger.info("Conversion Job Summary:")
    logger.info(f"Total PDFs found/extracted: {len(pdf_to_process)}")
    logger.info(f"Successful conversions/copies: {success_count}")
    logger.info(f"Total files skipped:           {skipped_count}")
    logger.info(f"Total files failed:            {failed_count}")
    logger.info("All operations completed.")
    logger.info(f"Logs have been saved to: {output_dir / 'converter.log'}")
    logger.info("=" * 60)

    # Post-processing print prompt
    interactive_print_prompt(output_dir)


if __name__ == "__main__":
    main()
